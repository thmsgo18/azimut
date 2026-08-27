"""Import Excel : relit un fichier généré par export_excel et réinjecte les
données dans la base — utile pour restaurer une sauvegarde ou fusionner.

Règles :
- toutes les écritures passent par les fonctions métier (jamais de SQL direct) ;
- une ligne déjà présente (doublon entreprise+poste ou entreprise+contact) est
  ignorée et signalée, jamais écrasée ;
- une ligne invalide (valeur hors liste, date impossible) est ignorée et
  signalée avec son numéro de ligne — le reste du fichier est importé ;
- les identifiants/mots de passe de portail ne figurent pas dans les exports :
  la sauvegarde complète reste le fichier suivi_candidatures.db.
"""

from pathlib import Path

import openpyxl

from candidatures import ajouter_candidature, verifier_doublon_candidature
from contacts import ajouter_contact, verifier_doublon_contact
from entreprises import ajouter_ou_recuperer_entreprise, lister_entreprises, modifier_entreprise
from exceptions import ConflitMiseAJour, ErreurSuivi, ValeurNonAutorisee
from valeurs import normaliser

# Correspondance en-tête de colonne -> champ de la base, par onglet.
COLONNES_SUIVI = {
    "Entreprise": "entreprise",
    "Date d'envoi": "date_envoi",
    "Poste / Intitulé": "poste",
    "Sous-domaine": "sous_domaine",
    "Lien de l'offre": "lien_offre",
    "Texte de l'offre": "texte_offre",
    "Type de candidature": "type_candidature",
    "Priorité": "priorite",
    "Statut": "statut",
    "Nb relances": "nb_relances",
    "Date de relance prévue": "date_relance_prevue",
    "Date de réponse": "date_reponse",
    "Date d'entretien": "date_entretien",
    "Date de début souhaitée": "date_debut_souhaitee",
    "Durée": "duree",
    "Gratification (€/mois)": "gratification",
    "Ville": "ville",
    "Mode de travail": "mode_travail",
    "Convention envoyée": "convention_envoyee",
    "Source": "source",
    "Notes": "notes",
    "Notes entretien": "notes_entretien",
}

COLONNES_ENTREPRISES = {
    "Nom": "nom",
    "Site web": "site_web",
    "Contexte / Actus": "contexte_actus",
    "Dernière recherche": "derniere_recherche",
}

COLONNES_CONTACTS = {
    "Entreprise": "entreprise",
    "Nom": "nom",
    "Poste": "poste",
    "Équipe": "equipe",
    "Email": "email",
    "Téléphone": "telephone",
    "LinkedIn": "linkedin",
    "Statut": "statut_contact",
    "Date de contact": "date_contact",
    "Source": "source",
    "Notes": "notes",
}


def _entetes(feuille):
    """{titre de colonne: index} pour la ligne 1 d'un onglet."""
    resultat = {}
    for i, cellule in enumerate(feuille[1], start=1):
        if cellule.value:
            resultat[str(cellule.value).strip()] = i
    return resultat


def _lignes(feuille, correspondance, premiere_ligne):
    """Itère (numéro de ligne, dict champ -> valeur brute) sur un onglet."""
    entetes = _entetes(feuille)
    colonnes = {
        champ: entetes[titre] for titre, champ in correspondance.items() if titre in entetes
    }
    for numero in range(premiere_ligne, feuille.max_row + 1):
        valeurs = {}
        for champ, index in colonnes.items():
            brut = feuille.cell(row=numero, column=index).value
            if brut is None or (isinstance(brut, str) and not brut.strip()):
                continue
            valeurs[champ] = brut.strip() if isinstance(brut, str) else brut
        if valeurs:
            yield numero, valeurs


def _nettoyer(valeurs):
    """Convertit les valeurs de cellules (dates datetime, nombres) en textes/ints."""
    import datetime

    nettoyees = {}
    for champ, valeur in valeurs.items():
        if isinstance(valeur, (datetime.datetime, datetime.date)):
            nettoyees[champ] = valeur.strftime("%Y-%m-%d")
        elif isinstance(valeur, float) and valeur.is_integer():
            nettoyees[champ] = int(valeur)
        else:
            nettoyees[champ] = valeur
    return nettoyees


def importer_excel(chemin_fichier, chemin_db=None):
    """Importe un fichier d'export Excel et retourne un rapport détaillé.

    Rapport : {"entreprises_ajoutees", "candidatures_ajoutees", "contacts_ajoutes",
    "ignores" (doublons, liste de textes), "erreurs" (liste de textes)}.
    """
    chemin = Path(chemin_fichier).expanduser()
    if not chemin.exists():
        raise ValeurNonAutorisee(f"Fichier introuvable : {chemin}")
    try:
        wb = openpyxl.load_workbook(chemin, data_only=True)
    except Exception:
        raise ValeurNonAutorisee(
            f"« {chemin.name} » n'est pas un classeur Excel lisible (.xlsx attendu)."
        )
    manquants = [
        onglet
        for onglet in ("Suivi candidatures", "Entreprises", "Contacts")
        if onglet not in wb.sheetnames
    ]
    if manquants:
        raise ValeurNonAutorisee(
            f"Onglet(s) manquant(s) : {', '.join(manquants)}. "
            "Ce fichier ne ressemble pas à un export de l'appli."
        )

    rapport = {
        "entreprises_ajoutees": 0,
        "candidatures_ajoutees": 0,
        "contacts_ajoutes": 0,
        "ignores": [],
        "erreurs": [],
    }
    noms_existants = {normaliser(e["nom"]) for e in lister_entreprises(chemin_db=chemin_db)}

    # 1. Entreprises (ligne 1 = en-têtes, données dès la ligne 2).
    for numero, valeurs in _lignes(wb["Entreprises"], COLONNES_ENTREPRISES, 2):
        valeurs = _nettoyer(valeurs)
        nom = valeurs.get("nom")
        if not nom:
            continue
        try:
            nouveau = normaliser(nom) not in noms_existants
            id_entreprise = ajouter_ou_recuperer_entreprise(
                nom,
                site_web=valeurs.get("site_web"),
                contexte_actus=valeurs.get("contexte_actus"),
                chemin_db=chemin_db,
            )
            if nouveau:
                rapport["entreprises_ajoutees"] += 1
                noms_existants.add(normaliser(nom))
                if valeurs.get("derniere_recherche"):
                    modifier_entreprise(
                        id_entreprise,
                        chemin_db=chemin_db,
                        derniere_recherche=valeurs["derniere_recherche"],
                    )
        except ConflitMiseAJour:
            rapport["ignores"].append(
                f"Entreprises ligne {numero} : « {nom} » existe déjà avec des infos "
                "différentes — rien n'a été écrasé."
            )
        except ErreurSuivi as erreur:
            rapport["erreurs"].append(f"Entreprises ligne {numero} : {erreur}")

    # 2. Candidatures (ligne 1 = en-têtes, ligne 2 = exemple, données dès la ligne 3).
    for numero, valeurs in _lignes(wb["Suivi candidatures"], COLONNES_SUIVI, 3):
        valeurs = _nettoyer(valeurs)
        entreprise = valeurs.pop("entreprise", None)
        poste = valeurs.pop("poste", None)
        if not entreprise or not poste:
            rapport["erreurs"].append(
                f"Candidatures ligne {numero} : entreprise ou poste manquant — ligne ignorée."
            )
            continue
        try:
            if verifier_doublon_candidature(entreprise, poste, chemin_db=chemin_db):
                rapport["ignores"].append(
                    f"Candidatures ligne {numero} : « {poste} » chez {entreprise} existe déjà."
                )
                continue
            ajouter_candidature(entreprise, poste, chemin_db=chemin_db, **valeurs)
            rapport["candidatures_ajoutees"] += 1
        except ErreurSuivi as erreur:
            rapport["erreurs"].append(f"Candidatures ligne {numero} : {erreur}")

    # 3. Contacts (données dès la ligne 2).
    for numero, valeurs in _lignes(wb["Contacts"], COLONNES_CONTACTS, 2):
        valeurs = _nettoyer(valeurs)
        entreprise = valeurs.pop("entreprise", None)
        nom = valeurs.pop("nom", None)
        if not entreprise or not nom:
            rapport["erreurs"].append(
                f"Contacts ligne {numero} : entreprise ou nom manquant — ligne ignorée."
            )
            continue
        try:
            if verifier_doublon_contact(entreprise, nom, chemin_db=chemin_db):
                rapport["ignores"].append(
                    f"Contacts ligne {numero} : {nom} ({entreprise}) existe déjà."
                )
                continue
            ajouter_contact(entreprise, nom, chemin_db=chemin_db, **valeurs)
            rapport["contacts_ajoutes"] += 1
        except ErreurSuivi as erreur:
            rapport["erreurs"].append(f"Contacts ligne {numero} : {erreur}")

    return rapport
