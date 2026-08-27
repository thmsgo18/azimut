"""Export Excel : régénère le fichier .xlsx complet depuis la base de données.

Reproduit le style du fichier « suivi candidatures.xlsx » existant (en-têtes
bleu foncé, Arial, bordures grises, ligne d'exemple jaune, tableau de bord à
formules) étendu aux 4 onglets du cahier des charges : Suivi candidatures,
Entreprises, Contacts, Tableau de bord. La base reste la seule source de
vérité : l'export peut être relancé à tout moment.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from candidatures import lister_candidatures
from contacts import lister_contacts
from entreprises import lister_entreprises
from valeurs import (
    CONVENTIONS,
    MODES_TRAVAIL,
    PRIORITES,
    SOURCES_CANDIDATURE,
    SOURCES_CONTACT,
    SOUS_DOMAINES,
    STATUTS,
    STATUTS_CONTACT,
    TYPES_CANDIDATURE,
)

# Palette et styles repris du fichier Excel existant.
BLEU_ENTETE = "FF1F4E78"
GRIS_BORDURE = "FFD9D9D9"
JAUNE_EXEMPLE = "FFFFF2CC"
GRIS_EXEMPLE = "FF7F7F7F"
BLEU_LIEN = "FF0563C1"

COULEURS_STATUT = {
    "À préparer": "FF595959",
    "Envoyée": "FF1F4E78",
    "Relancée": "FF7F6000",
    "Réponse reçue": "FF1F4E78",
    "Entretien": "FF833C00",
    "Refus": "FF9C0006",
    "Accepté": "FF375623",
}
COULEURS_PRIORITE = {"Haute": "FF9C0006", "Moyenne": "FF7F6000", "Basse": "FF595959"}

BORDURE = Border(
    left=Side(style="thin", color=GRIS_BORDURE),
    right=Side(style="thin", color=GRIS_BORDURE),
    top=Side(style="thin", color=GRIS_BORDURE),
    bottom=Side(style="thin", color=GRIS_BORDURE),
)
POLICE_ENTETE = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
FOND_ENTETE = PatternFill("solid", start_color=BLEU_ENTETE)
POLICE_DONNEES = Font(name="Arial", size=10)
POLICE_EXEMPLE = Font(name="Arial", size=10, italic=True, color=GRIS_EXEMPLE)
FOND_EXEMPLE = PatternFill("solid", start_color=JAUNE_EXEMPLE)
POLICE_LIEN = Font(name="Arial", size=10, color=BLEU_LIEN, underline="single")

LIGNE_MAX_MINI = 300


def _date_fr(iso):
    """AAAA-MM-JJ → JJ/MM/AAAA (les dates sont affichées comme dans le fichier existant)."""
    if not iso:
        return None
    try:
        annee, mois, jour = str(iso).split("-")
        return f"{jour}/{mois}/{annee}"
    except ValueError:
        return str(iso)


def _texte_cellule(valeur, maximum=32000):
    """Tronque les textes trop longs pour une cellule Excel (limite ~32 767 caractères)."""
    if valeur is None:
        return None
    texte = str(valeur)
    if len(texte) > maximum:
        return texte[:maximum] + " […]"
    return texte


def _preparer_onglet(ws, titres, largeurs, ligne_max, ligne_donnees):
    """Pose l'en-tête, les largeurs, les bordures et le volet figé d'un onglet de données."""
    for i, (titre, largeur) in enumerate(zip(titres, largeurs), start=1):
        cellule = ws.cell(row=1, column=i, value=titre)
        cellule.font = POLICE_ENTETE
        cellule.fill = FOND_ENTETE
        cellule.border = BORDURE
        cellule.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.row_dimensions[1].height = 31.5
    for ligne in range(2, ligne_max + 1):
        for col in range(1, len(titres) + 1):
            cellule = ws.cell(row=ligne, column=col)
            cellule.font = POLICE_DONNEES
            cellule.border = BORDURE
            cellule.alignment = Alignment(vertical="center")
    ws.freeze_panes = f"A{ligne_donnees}"


def _ajouter_validation(ws, colonne, valeurs, premiere_ligne, ligne_max):
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(valeurs) + '"',
        allow_blank=True,
        showErrorMessage=True,
    )
    dv.error = "Valeur non autorisée - choisir dans la liste déroulante."
    dv.errorTitle = "Valeur non autorisée"
    ws.add_data_validation(dv)
    dv.add(f"{colonne}{premiere_ligne}:{colonne}{ligne_max}")


def _onglet_suivi(wb, liste, ligne_max):
    ws = wb.active
    ws.title = "Suivi candidatures"
    titres = [
        "Entreprise",
        "Date d'envoi",
        "Poste / Intitulé",
        "Sous-domaine",
        "Lien de l'offre",
        "Texte de l'offre",
        "Type de candidature",
        "Priorité",
        "Statut",
        "Nb relances",
        "Date de relance prévue",
        "Date de réponse",
        "Date d'entretien",
        "Date de début souhaitée",
        "Durée",
        "Gratification (€/mois)",
        "Ville",
        "Mode de travail",
        "Convention envoyée",
        "Source",
        "Notes",
        "Notes entretien",
        "Fiche entreprise",
        "Contacts",
    ]
    largeurs = [20, 14, 26, 24, 30, 18, 20, 12, 16, 12, 20, 16, 16, 20, 14, 18, 16, 16, 16, 14, 30, 24, 16, 14]
    _preparer_onglet(ws, titres, largeurs, ligne_max, ligne_donnees=3)

    # Ligne 2 : exemple (fond jaune, italique), comme dans le fichier existant.
    exemple = [
        "Ex: Mistral AI",
        "08/09/2026",
        "Stage - Agents LLM autonomes",
        "Orchestration multi-agents",
        "https://exemple.com/offre",
        "Texte intégral de l'offre (archive)",
        "Offre publiée",
        "Haute",
        "Envoyée",
        0,
        "15/09/2026",
        None,
        None,
        "02/02/2027",
        "5 mois",
        1400,
        "Paris",
        "Hybride",
        "Non",
        "LinkedIn",
        "Relancer si pas de réponse sous 2 semaines",
        None,
        None,
        None,
    ]
    for i, valeur in enumerate(exemple, start=1):
        cellule = ws.cell(row=2, column=i, value=valeur)
        cellule.font = POLICE_EXEMPLE
        cellule.fill = FOND_EXEMPLE

    # Données réelles à partir de la ligne 3.
    for i, cand in enumerate(liste):
        ligne = 3 + i
        valeurs_ligne = [
            cand["entreprise"],
            _date_fr(cand["date_envoi"]),
            cand["poste"],
            cand["sous_domaine"],
            cand["lien_offre"],
            _texte_cellule(cand["texte_offre"]),
            cand["type_candidature"],
            cand["priorite"],
            cand["statut"],
            cand["nb_relances"],
            _date_fr(cand["date_relance_prevue"]),
            _date_fr(cand["date_reponse"]),
            _date_fr(cand["date_entretien"]),
            _date_fr(cand["date_debut_souhaitee"]),
            cand["duree"],
            cand["gratification"],
            cand["ville"],
            cand["mode_travail"],
            cand["convention_envoyee"],
            cand["source"],
            _texte_cellule(cand["notes"]),
            _texte_cellule(cand.get("notes_entretien")),
        ]
        for col, valeur in enumerate(valeurs_ligne, start=1):
            ws.cell(row=ligne, column=col, value=valeur)
        if cand["lien_offre"]:
            cellule = ws.cell(row=ligne, column=5)
            cellule.hyperlink = cand["lien_offre"]
            cellule.font = POLICE_LIEN
        # Liens HYPERLINK + MATCH vers les onglets Entreprises et Contacts.
        fiche = ws.cell(
            row=ligne,
            column=23,
            value=f'=HYPERLINK("#Entreprises!A"&MATCH($A{ligne},Entreprises!$A:$A,0),"→ Entreprise")',
        )
        fiche.font = POLICE_LIEN
        lien_contacts = ws.cell(
            row=ligne,
            column=24,
            value=f'=IFERROR(HYPERLINK("#Contacts!A"&MATCH($A{ligne},Contacts!$A:$A,0),"→ Contacts"),"(aucun)")',
        )
        lien_contacts.font = POLICE_LIEN

    # Listes déroulantes sur les colonnes à valeurs autorisées.
    for colonne, valeurs in [
        ("D", SOUS_DOMAINES),
        ("G", TYPES_CANDIDATURE),
        ("H", PRIORITES),
        ("I", STATUTS),
        ("R", MODES_TRAVAIL),
        ("S", CONVENTIONS),
        ("T", SOURCES_CANDIDATURE),
    ]:
        _ajouter_validation(ws, colonne, valeurs, 3, ligne_max)

    # Mise en forme conditionnelle (couleur de police) sur Statut et Priorité.
    for statut, couleur in COULEURS_STATUT.items():
        ws.conditional_formatting.add(
            f"I3:I{ligne_max}",
            CellIsRule(operator="equal", formula=[f'"{statut}"'], font=Font(color=couleur)),
        )
    for priorite, couleur in COULEURS_PRIORITE.items():
        ws.conditional_formatting.add(
            f"H3:H{ligne_max}",
            CellIsRule(operator="equal", formula=[f'"{priorite}"'], font=Font(color=couleur)),
        )
    return ws


def _onglet_entreprises(wb, liste, ligne_max, ligne_max_suivi):
    ws = wb.create_sheet("Entreprises")
    titres = ["Nom", "Site web", "Contexte / Actus", "Dernière recherche", "Nb candidatures", "Candidatures"]
    largeurs = [26, 32, 60, 18, 16, 18]
    _preparer_onglet(ws, titres, largeurs, ligne_max, ligne_donnees=2)
    for i, ent in enumerate(liste):
        ligne = 2 + i
        ws.cell(row=ligne, column=1, value=ent["nom"])
        if ent["site_web"]:
            cellule = ws.cell(row=ligne, column=2, value=ent["site_web"])
            cellule.hyperlink = ent["site_web"]
            cellule.font = POLICE_LIEN
        cellule = ws.cell(row=ligne, column=3, value=_texte_cellule(ent["contexte_actus"]))
        cellule.alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=ligne, column=4, value=_date_fr(ent["derniere_recherche"]))
        ws.cell(
            row=ligne,
            column=5,
            value=f"=COUNTIF('Suivi candidatures'!$A$3:$A${ligne_max_suivi},$A{ligne})",
        ).alignment = Alignment(horizontal="center", vertical="center")
        lien = ws.cell(
            row=ligne,
            column=6,
            value=(
                "=IFERROR(HYPERLINK(\"#'Suivi candidatures'!A\"&"
                f"MATCH($A{ligne},'Suivi candidatures'!$A:$A,0),\"→ Candidatures\"),\"(aucune)\")"
            ),
        )
        lien.font = POLICE_LIEN
    return ws


def _onglet_contacts(wb, liste, ligne_max):
    ws = wb.create_sheet("Contacts")
    titres = [
        "Entreprise",
        "Nom",
        "Poste",
        "Équipe",
        "Email",
        "Téléphone",
        "LinkedIn",
        "Statut",
        "Date de contact",
        "Source",
        "Notes",
        "Offre associée",
    ]
    largeurs = [20, 20, 24, 18, 24, 16, 30, 16, 16, 24, 30, 14]
    _preparer_onglet(ws, titres, largeurs, ligne_max, ligne_donnees=2)
    for i, contact in enumerate(liste):
        ligne = 2 + i
        valeurs_ligne = [
            contact["entreprise"],
            contact["nom"],
            contact["poste"],
            contact["equipe"],
            contact["email"],
            contact["telephone"],
            contact["linkedin"],
            contact["statut_contact"],
            _date_fr(contact["date_contact"]),
            contact["source"],
            _texte_cellule(contact["notes"]),
        ]
        for col, valeur in enumerate(valeurs_ligne, start=1):
            ws.cell(row=ligne, column=col, value=valeur)
        lien = ws.cell(
            row=ligne,
            column=12,
            value=(
                "=IFERROR(HYPERLINK(\"#'Suivi candidatures'!A\"&"
                f"MATCH($A{ligne},'Suivi candidatures'!$A:$A,0),\"→ Offre\"),\"(aucune)\")"
            ),
        )
        lien.font = POLICE_LIEN
    for colonne, valeurs in [
        ("H", STATUTS_CONTACT),
        ("J", SOURCES_CONTACT),
    ]:
        _ajouter_validation(ws, colonne, valeurs, 2, ligne_max)
    return ws


def _onglet_tableau_de_bord(wb, ligne_max_suivi, ligne_max_contacts):
    ws = wb.create_sheet("Tableau de bord")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 4
    ws.column_dimensions["D"].width = 70

    titre = ws["A1"]
    titre.value = "Tableau de bord - candidatures stage"
    titre.font = Font(name="Arial", size=14, bold=True, color=BLEU_ENTETE)
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 17.35

    def entete_section(ligne, libelle):
        for colonne, valeur in (("A", libelle), ("B", "Nombre")):
            cellule = ws[f"{colonne}{ligne}"]
            cellule.value = valeur
            cellule.font = POLICE_ENTETE
            cellule.fill = FOND_ENTETE

    def ligne_compteur(ligne, libelle, formule, gras_libelle=False):
        cellule_a = ws[f"A{ligne}"]
        cellule_a.value = libelle
        cellule_a.font = Font(name="Arial", size=11 if gras_libelle else 10, bold=gras_libelle)
        cellule_a.border = BORDURE
        cellule_b = ws[f"B{ligne}"]
        cellule_b.value = formule
        cellule_b.font = Font(name="Arial", size=11 if gras_libelle else 10, bold=True)
        cellule_b.border = BORDURE
        cellule_b.alignment = Alignment(horizontal="center")
        return cellule_b

    plage_statut = f"'Suivi candidatures'!$I$3:$I${ligne_max_suivi}"
    entete_section(3, "Statut")
    for i, statut in enumerate(STATUTS):
        ligne_compteur(4 + i, statut, f'=COUNTIF({plage_statut},"{statut}")')
    ligne_compteur(
        12,
        "Total candidatures",
        f"=COUNTA('Suivi candidatures'!$A$3:$A${ligne_max_suivi})",
        gras_libelle=True,
    )
    taux = ligne_compteur(13, "Taux de réponse", "=IFERROR((B7+B8+B9+B10)/B12,0)", gras_libelle=True)
    taux.number_format = "0%"

    plage_domaine = f"'Suivi candidatures'!$D$3:$D${ligne_max_suivi}"
    entete_section(15, "Sous-domaine")
    for i, domaine in enumerate(SOUS_DOMAINES):
        ligne_compteur(16 + i, domaine, f'=COUNTIF({plage_domaine},"{domaine}")')

    plage_contacts = f"Contacts!$H$2:$H${ligne_max_contacts}"
    entete_section(24, "Statut des contacts")
    for i, statut in enumerate(STATUTS_CONTACT):
        ligne_compteur(25 + i, statut, f'=COUNTIF({plage_contacts},"{statut}")')
    ligne_compteur(
        30, "Total contacts", f"=COUNTA(Contacts!$B$2:$B${ligne_max_contacts})", gras_libelle=True
    )

    conseils = [
        "1. Ce fichier est généré automatiquement depuis la base de données "
        "(suivi_candidatures.db) : la base reste la seule source de vérité.",
        "2. Ne pas le modifier à la main - toute modification sera perdue au prochain export.",
        "3. Pour ajouter ou corriger une ligne : passer par la CLI (python cli.py ...) "
        "ou par Claude Code, puis relancer « export excel ».",
        "4. La ligne 2 (fond jaune, italique) de « Suivi candidatures » est un exemple de format.",
        "5. Les compteurs ci-contre se mettent à jour automatiquement via des formules.",
        "6. Les colonnes « Fiche entreprise », « Contacts », « Candidatures » et « Offre associée » "
        "sont des liens cliquables entre onglets.",
    ]
    ws["D3"].value = "Comment utiliser ce fichier"
    ws["D3"].font = Font(name="Arial", size=11, bold=True, color=BLEU_ENTETE)
    for i, texte in enumerate(conseils):
        cellule = ws[f"D{4 + i}"]
        cellule.value = texte
        cellule.font = Font(name="Arial", size=10)
        cellule.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[4 + i].height = 23.85
    return ws


def exporter_excel(chemin_sortie, chemin_db=None):
    """Génère le fichier .xlsx complet depuis la base et retourne son chemin."""
    liste_candidatures = lister_candidatures(chemin_db=chemin_db)
    liste_entreprises = lister_entreprises(chemin_db=chemin_db)
    liste_contacts = lister_contacts(chemin_db=chemin_db)

    ligne_max_suivi = max(LIGNE_MAX_MINI, len(liste_candidatures) + 2)
    ligne_max_entreprises = max(LIGNE_MAX_MINI, len(liste_entreprises) + 1)
    ligne_max_contacts = max(LIGNE_MAX_MINI, len(liste_contacts) + 1)

    wb = Workbook()
    _onglet_suivi(wb, liste_candidatures, ligne_max_suivi)
    _onglet_entreprises(wb, liste_entreprises, ligne_max_entreprises, ligne_max_suivi)
    _onglet_contacts(wb, liste_contacts, ligne_max_contacts)
    _onglet_tableau_de_bord(wb, ligne_max_suivi, ligne_max_contacts)

    chemin = Path(chemin_sortie).expanduser()
    chemin.parent.mkdir(parents=True, exist_ok=True)
    wb.save(chemin)
    return str(chemin)
