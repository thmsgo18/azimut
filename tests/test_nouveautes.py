"""Tests des nouveautés : journal, documents, réglages, recherche globale,
statistiques avancées, agenda (.ics), sauvegardes, notes d'entretien, agent."""

import io
import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import db
from exceptions import ChampInconnu, ValeurNonAutorisee


class TestModulesNouveautes(unittest.TestCase):
    def setUp(self):
        self.dossier = tempfile.TemporaryDirectory()
        self.chemin_db = str(Path(self.dossier.name) / "test.db")
        db.initialiser_base(self.chemin_db)

    def tearDown(self):
        self.dossier.cleanup()

    def _candidature(self, entreprise="AgentikCo", poste="Stage agents IA", **champs):
        from candidatures import ajouter_candidature

        return ajouter_candidature(entreprise, poste, chemin_db=self.chemin_db, **champs)

    # --- journal (timeline) ---

    def test_journal_alimente_automatiquement(self):
        from candidatures import modifier_candidature
        from evenements import lister_evenements

        numero = self._candidature(statut="Envoyée")
        modifier_candidature(numero, chemin_db=self.chemin_db, statut="Relancée", nb_relances=1)
        modifier_candidature(numero, chemin_db=self.chemin_db, date_reponse="2026-08-25")
        modifier_candidature(numero, chemin_db=self.chemin_db, date_entretien="05/09/2026")
        types = [e["type_evenement"] for e in lister_evenements(numero, chemin_db=self.chemin_db)]
        self.assertEqual(len(types), 5)
        self.assertIn("creation", types)
        self.assertIn("statut", types)
        self.assertIn("relance", types)
        self.assertIn("reponse", types)
        self.assertIn("entretien", types)

    def test_journal_pas_devenement_sans_changement(self):
        from candidatures import modifier_candidature
        from evenements import lister_evenements

        numero = self._candidature(statut="Envoyée")
        modifier_candidature(numero, chemin_db=self.chemin_db, ville="Paris")
        evenements_liste = lister_evenements(numero, chemin_db=self.chemin_db)
        self.assertEqual(len(evenements_liste), 1)  # seulement la création

    # --- documents ---

    def test_documents_cycle_de_vie(self):
        import documents

        numero = self._candidature()
        id_doc = documents.ajouter_document(
            numero, "CV Thomas v3.pdf", b"%PDF-1.4 faux contenu",
            type_document="cv", chemin_db=self.chemin_db,
        )
        liste = documents.lister_documents(chemin_db=self.chemin_db)
        self.assertEqual(liste[0]["type_document"], "CV")
        self.assertEqual(liste[0]["entreprise"], "AgentikCo")
        chemin = Path(documents.recuperer_document(id_doc, chemin_db=self.chemin_db)["chemin_absolu"])
        self.assertTrue(chemin.exists())
        documents.supprimer_document(id_doc, chemin_db=self.chemin_db)
        self.assertFalse(chemin.exists())
        self.assertEqual(documents.lister_documents(chemin_db=self.chemin_db), [])

    def test_document_type_invalide_ou_vide(self):
        import documents

        numero = self._candidature()
        with self.assertRaises(ValeurNonAutorisee):
            documents.ajouter_document(numero, "x.pdf", b"abc", type_document="Selfie",
                                       chemin_db=self.chemin_db)
        with self.assertRaises(ValeurNonAutorisee):
            documents.ajouter_document(numero, "x.pdf", b"", chemin_db=self.chemin_db)

    def test_suppression_candidature_nettoie_documents_et_journal(self):
        import documents
        from candidatures import supprimer_candidature
        from evenements import lister_evenements

        numero = self._candidature()
        id_doc = documents.ajouter_document(numero, "cv.pdf", b"abc", chemin_db=self.chemin_db)
        chemin = Path(documents.recuperer_document(id_doc, chemin_db=self.chemin_db)["chemin_absolu"])
        supprimer_candidature(numero, chemin_db=self.chemin_db)
        self.assertEqual(documents.lister_documents(chemin_db=self.chemin_db), [])
        self.assertEqual(lister_evenements(numero, chemin_db=self.chemin_db), [])
        self.assertFalse(chemin.exists())

    # --- réglages ---

    def test_reglages_masquage_et_defauts(self):
        import reglages

        etat = reglages.etat_reglages(chemin_db=self.chemin_db)
        self.assertFalse(etat["cle_api_definie"])
        self.assertEqual(etat["modele_ia"], "claude-opus-5")
        reglages.definir_reglage("cle_api", "sk-ant-api-tres-secrete-1234", chemin_db=self.chemin_db)
        etat = reglages.etat_reglages(chemin_db=self.chemin_db)
        self.assertTrue(etat["cle_api_definie"])
        self.assertNotIn("secrete", etat["cle_api_masquee"])
        self.assertTrue(etat["cle_api_masquee"].endswith("1234"))
        reglages.definir_reglage("cle_api", "", chemin_db=self.chemin_db)
        self.assertFalse(reglages.etat_reglages(chemin_db=self.chemin_db)["cle_api_definie"])
        with self.assertRaises(ChampInconnu):
            reglages.definir_reglage("mot_de_passe_maitre", "x", chemin_db=self.chemin_db)

    # --- recherche globale ---

    def test_recherche_types_et_accents(self):
        from contacts import ajouter_contact
        from recherche import rechercher

        self._candidature(notes="Équipe très réactive, poste orienté évaluation d'agents")
        ajouter_contact("AgentikCo", "Éléonore Petit", poste="Lead AI", chemin_db=self.chemin_db)
        resultats = rechercher("eleonore", chemin_db=self.chemin_db)
        self.assertEqual(len(resultats["contacts"]), 1)
        self.assertEqual(resultats["contacts"][0]["champs_trouves"], ["Nom"])
        resultats = rechercher("EVALUATION", chemin_db=self.chemin_db)
        self.assertEqual(len(resultats["candidatures"]), 1)
        self.assertIn("Notes", resultats["candidatures"][0]["champs_trouves"])
        self.assertIn("évaluation", resultats["candidatures"][0]["extrait"])
        resultats = rechercher("agentik", chemin_db=self.chemin_db)
        self.assertEqual(len(resultats["entreprises"]), 1)
        self.assertEqual(rechercher("", chemin_db=self.chemin_db),
                         {"candidatures": [], "entreprises": [], "contacts": []})

    # --- statistiques avancées ---

    def test_stats_avancees(self):
        from statistiques import stats_avancees

        self._candidature(statut="Refus", date_envoi="2026-08-01", date_reponse="2026-08-11",
                          source="LinkedIn")
        self._candidature(poste="Stage RAG", statut="Entretien", date_envoi="2026-08-05",
                          date_reponse="2026-08-10", date_entretien="2026-08-20", source="LinkedIn")
        self._candidature(entreprise="Mistral AI", poste="Stage évals", statut="Envoyée",
                          date_envoi="2026-08-20", source="Réseau")
        stats = stats_avancees(chemin_db=self.chemin_db)
        entonnoir = {e["etape"]: e for e in stats["entonnoir"]}
        self.assertEqual(entonnoir["Envoyées"]["nombre"], 3)
        self.assertEqual(entonnoir["Réponses"]["nombre"], 2)
        self.assertEqual(entonnoir["Entretiens"]["nombre"], 1)
        self.assertEqual(stats["delai_moyen_reponse"], 7.5)  # (10 + 5) / 2
        linkedin = next(s for s in stats["par_source"] if s["source"] == "LinkedIn")
        self.assertEqual(linkedin["envoyees"], 2)
        self.assertEqual(linkedin["taux"], 100)

    # --- agenda ---

    def test_agenda_et_ics(self):
        from agenda import generer_ics, lister_echeances

        self._candidature(date_relance_prevue="2099-01-10", date_entretien="2099-01-15")
        echeances = lister_echeances(chemin_db=self.chemin_db)
        self.assertEqual([e["type"] for e in echeances], ["relance", "entretien"])
        ics = generer_ics(chemin_db=self.chemin_db)
        self.assertIn("BEGIN:VCALENDAR", ics)
        self.assertEqual(ics.count("BEGIN:VEVENT"), 2)
        self.assertIn("DTSTART;VALUE=DATE:20990115", ics)
        self.assertIn("Entretien - AgentikCo", ics.replace("\\,", ","))
        self.assertIn("BEGIN:VALARM", ics)

    # --- sauvegardes ---

    def test_sauvegarde_et_rotation(self):
        import sauvegarde

        self._candidature()
        dossier_origine = sauvegarde.DOSSIER_SAUVEGARDES_DEFAUT
        sauvegarde.DOSSIER_SAUVEGARDES_DEFAUT = Path(self.dossier.name) / "sauvegardes"
        try:
            chemins = set()
            for _ in range(4):
                chemin = sauvegarde.sauvegarder_base(chemin_db=self.chemin_db, garder=3)
                self.assertIsNotNone(chemin)
                chemins.add(chemin)
            restantes = list(sauvegarde.DOSSIER_SAUVEGARDES_DEFAUT.glob("*.db"))
            self.assertLessEqual(len(restantes), 3)
            absente = sauvegarde.sauvegarder_base(
                chemin_db=str(Path(self.dossier.name) / "inexistante.db")
            )
            self.assertIsNone(absente)
        finally:
            sauvegarde.DOSSIER_SAUVEGARDES_DEFAUT = dossier_origine

    def test_sauvegarde_respecte_dossier_donnees_choisi(self):
        import reglages
        import sauvegarde

        self._candidature()
        personnalise = Path(self.dossier.name) / "mon-dossier-perso"
        reglages.definir_dossier_donnees(str(personnalise), chemin_db=self.chemin_db)
        chemin = sauvegarde.sauvegarder_base(chemin_db=self.chemin_db)
        self.assertIsNotNone(chemin)
        self.assertTrue(str(personnalise) in chemin)
        self.assertTrue((personnalise / "sauvegardes").exists())

    # --- notes d'entretien : fiche + export/import ---

    def test_notes_entretien_fiche_et_export(self):
        import openpyxl

        from candidatures import modifier_candidature
        from entretien import generer_fiche_entretien
        from export_excel import exporter_excel
        from import_excel import importer_excel

        numero = self._candidature(statut="Entretien")
        modifier_candidature(numero, chemin_db=self.chemin_db,
                             notes_entretien="Question posée : architecture des évals.")
        fiche = generer_fiche_entretien(numero, chemin_db=self.chemin_db)
        self.assertIn("Notes d'entretien : Question posée", fiche)
        self.assertIn("## Journal", fiche)
        self.assertIn("Candidature créée", fiche)

        chemin_xlsx = str(Path(self.dossier.name) / "export.xlsx")
        exporter_excel(chemin_xlsx, chemin_db=self.chemin_db)
        ws = openpyxl.load_workbook(chemin_xlsx)["Suivi candidatures"]
        self.assertEqual(ws["V1"].value, "Notes entretien")
        self.assertIn("Question posée", ws["V3"].value)

        cible = str(Path(self.dossier.name) / "cible.db")
        db.initialiser_base(cible)
        importer_excel(chemin_xlsx, chemin_db=cible)
        from candidatures import lister_candidatures

        self.assertIn("Question posée", lister_candidatures(chemin_db=cible)[0]["notes_entretien"])


class TestApiNouveautes(unittest.TestCase):
    def setUp(self):
        self.dossier = tempfile.TemporaryDirectory()
        self.chemin_origine = db.CHEMIN_DB
        db.CHEMIN_DB = Path(self.dossier.name) / "test.db"
        db.initialiser_base()
        from serveur import app

        app.config["TESTING"] = True
        self.client = app.test_client()

    def tearDown(self):
        db.CHEMIN_DB = self.chemin_origine
        self.dossier.cleanup()

    def _ajouter(self, **surcharge):
        donnees = {"entreprise": "AgentikCo", "poste": "Stage agents IA", "statut": "Envoyée"}
        donnees.update(surcharge)
        return self.client.post("/api/candidatures", json=donnees).get_json()["id"]

    def test_evenements_endpoint(self):
        numero = self._ajouter()
        self.client.patch(f"/api/candidatures/{numero}", json={"statut": "Entretien"})
        journal = self.client.get(f"/api/candidatures/{numero}/evenements").get_json()
        self.assertEqual(len(journal), 2)
        self.assertEqual(self.client.get("/api/candidatures/999/evenements").status_code, 404)

    def test_documents_endpoints(self):
        numero = self._ajouter()
        envoi = self.client.post(
            f"/api/candidatures/{numero}/documents",
            data={"fichier": (io.BytesIO(b"faux pdf"), "CV.pdf"), "type": "CV"},
            content_type="multipart/form-data",
        )
        self.assertEqual(envoi.status_code, 201)
        id_doc = envoi.get_json()["id"]
        liste = self.client.get("/api/documents").get_json()
        self.assertEqual(liste[0]["nom_fichier"], "CV.pdf")
        telechargement = self.client.get(f"/api/documents/{id_doc}/telecharger")
        self.assertEqual(telechargement.status_code, 200)
        self.assertEqual(telechargement.data, b"faux pdf")
        self.assertEqual(self.client.delete(f"/api/documents/{id_doc}").status_code, 200)
        sans_fichier = self.client.post(f"/api/candidatures/{numero}/documents")
        self.assertEqual(sans_fichier.status_code, 400)

    def test_reglages_endpoint_masque_la_cle(self):
        reponse = self.client.post(
            "/api/reglages", json={"cle_api": "sk-ant-tres-secret-9876", "modele_ia": "claude-sonnet-5"}
        )
        etat = reponse.get_json()
        self.assertTrue(etat["cle_api_definie"])
        self.assertNotIn("tres-secret", str(etat))
        self.assertEqual(etat["modele_ia"], "claude-sonnet-5")
        inconnu = self.client.post("/api/reglages", json={"cle_api_bis": "x"})
        self.assertEqual(inconnu.status_code, 200)  # clé inconnue simplement ignorée

    def test_agent_sans_cle_renvoie_400(self):
        reponse = self.client.post("/api/agent/analyser", json={"texte": "Stage agents IA à Paris"})
        self.assertEqual(reponse.status_code, 400)
        self.assertIn("Réglages", reponse.get_json()["erreur"])
        test = self.client.post("/api/agent/tester")
        self.assertEqual(test.status_code, 400)

    def test_recherche_endpoint(self):
        self._ajouter(ville="Paris")
        resultats = self.client.get("/api/recherche?q=paris").get_json()
        self.assertEqual(len(resultats["candidatures"]), 1)
        vide = self.client.get("/api/recherche").get_json()
        self.assertEqual(vide["candidatures"], [])

    def test_stats_agenda_ics_endpoints(self):
        self._ajouter(date_entretien="2099-03-01")
        self.assertEqual(self.client.get("/api/stats/avancees").status_code, 200)
        agenda_liste = self.client.get("/api/agenda").get_json()
        self.assertEqual(agenda_liste[0]["type"], "entretien")
        ics = self.client.get("/api/agenda/ics")
        self.assertEqual(ics.status_code, 200)
        self.assertIn("text/calendar", ics.headers["Content-Type"])

    def test_notes_entretien_par_api(self):
        numero = self._ajouter()
        self.client.patch(f"/api/candidatures/{numero}", json={"notes_entretien": "Très bon échange."})
        relu = self.client.get(f"/api/candidatures/{numero}").get_json()
        self.assertEqual(relu["notes_entretien"], "Très bon échange.")


if __name__ == "__main__":
    unittest.main()
