"""Tests de l'import Excel : aller-retour export → import, doublons, fichiers invalides."""

import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

import db
from candidatures import ajouter_candidature, lister_candidatures
from contacts import ajouter_contact, lister_contacts
from entreprises import ajouter_ou_recuperer_entreprise, lister_entreprises
from exceptions import ValeurNonAutorisee
from export_excel import exporter_excel
from import_excel import importer_excel


class TestImportExcel(unittest.TestCase):
    def setUp(self):
        self.dossier = tempfile.TemporaryDirectory()
        self.source = str(Path(self.dossier.name) / "source.db")
        self.cible = str(Path(self.dossier.name) / "cible.db")
        db.initialiser_base(self.source)
        db.initialiser_base(self.cible)

    def tearDown(self):
        self.dossier.cleanup()

    def _peupler_source(self):
        ajouter_ou_recuperer_entreprise(
            "AgentikCo",
            site_web="https://agentik.co",
            contexte_actus="Série A en 2026.",
            chemin_db=self.source,
        )
        ajouter_candidature(
            "AgentikCo",
            "Stage agents IA",
            date_envoi="2026-08-20",
            statut="Entretien",
            priorite="Haute",
            sous_domaine="Orchestration multi-agents",
            gratification=1400,
            nb_relances=2,
            ville="Paris",
            texte_offre="Concevoir des agents multi-étapes.",
            chemin_db=self.source,
        )
        ajouter_candidature("Mistral AI", "Stage RAG", chemin_db=self.source)
        ajouter_contact(
            "AgentikCo",
            "Marie Petit",
            poste="Lead AI",
            type_contact="Email",
            valeur_contact="marie@agentik.co",
            statut_contact="Répondu",
            date_contact="2026-08-12",
            chemin_db=self.source,
        )

    def _exporter(self):
        chemin = str(Path(self.dossier.name) / "export.xlsx")
        exporter_excel(chemin, chemin_db=self.source)
        return chemin

    def test_aller_retour_complet(self):
        """Un export réimporté dans une base vierge restitue toutes les données."""
        self._peupler_source()
        rapport = importer_excel(self._exporter(), chemin_db=self.cible)
        self.assertEqual(rapport["candidatures_ajoutees"], 2)
        self.assertEqual(rapport["contacts_ajoutes"], 1)
        self.assertEqual(rapport["entreprises_ajoutees"], 2)
        self.assertEqual(rapport["erreurs"], [])
        self.assertEqual(rapport["ignores"], [])

        cands = lister_candidatures(chemin_db=self.cible)
        self.assertEqual(len(cands), 2)
        agentik = next(c for c in cands if c["entreprise"] == "AgentikCo")
        self.assertEqual(agentik["date_envoi"], "2026-08-20")  # JJ/MM/AAAA → ISO
        self.assertEqual(agentik["statut"], "Entretien")
        self.assertEqual(agentik["gratification"], 1400)
        self.assertEqual(agentik["nb_relances"], 2)
        self.assertEqual(agentik["texte_offre"], "Concevoir des agents multi-étapes.")

        contacts_cible = lister_contacts(chemin_db=self.cible)
        self.assertEqual(contacts_cible[0]["nom"], "Marie Petit")
        self.assertEqual(contacts_cible[0]["date_contact"], "2026-08-12")
        self.assertEqual(contacts_cible[0]["statut_contact"], "Répondu")

        entreprises_cible = lister_entreprises(chemin_db=self.cible)
        agentik_ent = next(e for e in entreprises_cible if e["nom"] == "AgentikCo")
        self.assertEqual(agentik_ent["contexte_actus"], "Série A en 2026.")

    def test_import_idempotent(self):
        """Importer deux fois le même fichier n'ajoute rien et signale les doublons."""
        self._peupler_source()
        chemin = self._exporter()
        importer_excel(chemin, chemin_db=self.cible)
        rapport = importer_excel(chemin, chemin_db=self.cible)
        self.assertEqual(rapport["candidatures_ajoutees"], 0)
        self.assertEqual(rapport["contacts_ajoutes"], 0)
        self.assertEqual(rapport["entreprises_ajoutees"], 0)
        self.assertEqual(len(rapport["ignores"]), 3)  # 2 candidatures + 1 contact
        self.assertEqual(len(lister_candidatures(chemin_db=self.cible)), 2)

    def test_ligne_exemple_ignoree(self):
        """La ligne 2 (exemple jaune) de l'export n'est jamais importée."""
        self._peupler_source()
        importer_excel(self._exporter(), chemin_db=self.cible)
        noms = {e["nom"] for e in lister_entreprises(chemin_db=self.cible)}
        self.assertNotIn("Ex: Mistral AI", noms)

    def test_ligne_invalide_signalee_sans_bloquer(self):
        """Une valeur hors liste bloque sa ligne, pas le reste de l'import."""
        self._peupler_source()
        chemin = self._exporter()
        wb = openpyxl.load_workbook(chemin)
        ws = wb["Suivi candidatures"]
        ws.cell(row=5, column=1, value="NouvelleCo")
        ws.cell(row=5, column=3, value="Stage test")
        ws.cell(row=5, column=9, value="Statut inventé")
        wb.save(chemin)
        rapport = importer_excel(chemin, chemin_db=self.cible)
        self.assertEqual(rapport["candidatures_ajoutees"], 2)
        self.assertEqual(len(rapport["erreurs"]), 1)
        self.assertIn("ligne 5", rapport["erreurs"][0])
        self.assertIn("Statut inventé", rapport["erreurs"][0])

    def test_fichier_inexistant_ou_invalide(self):
        with self.assertRaises(ValeurNonAutorisee):
            importer_excel(Path(self.dossier.name) / "absent.xlsx", chemin_db=self.cible)
        mauvais = Path(self.dossier.name) / "mauvais.xlsx"
        mauvais.write_text("ceci n'est pas un classeur")
        with self.assertRaises(ValeurNonAutorisee):
            importer_excel(mauvais, chemin_db=self.cible)

    def test_classeur_sans_les_bons_onglets(self):
        etranger = Path(self.dossier.name) / "etranger.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Feuille quelconque"
        wb.save(etranger)
        with self.assertRaises(ValeurNonAutorisee) as contexte:
            importer_excel(etranger, chemin_db=self.cible)
        self.assertIn("Onglet(s) manquant(s)", str(contexte.exception))

    def test_import_fusionne_sans_ecraser(self):
        """Importer dans une base déjà remplie ajoute le nouveau, préserve l'existant."""
        self._peupler_source()
        chemin = self._exporter()
        # La cible contient déjà AgentikCo avec un autre contexte.
        ajouter_ou_recuperer_entreprise(
            "AgentikCo", contexte_actus="Contexte local à préserver.", chemin_db=self.cible
        )
        rapport = importer_excel(chemin, chemin_db=self.cible)
        agentik = next(
            e for e in lister_entreprises(chemin_db=self.cible) if e["nom"] == "AgentikCo"
        )
        self.assertEqual(agentik["contexte_actus"], "Contexte local à préserver.")
        self.assertTrue(any("AgentikCo" in ligne for ligne in rapport["ignores"]))
        self.assertEqual(rapport["candidatures_ajoutees"], 2)

    def test_export_sans_mot_de_passe_portail(self):
        """Les identifiants de portail ne sortent jamais dans un export Excel."""
        self._peupler_source()
        from candidatures import lister_candidatures as lister
        from candidatures import modifier_candidature

        numero = lister(chemin_db=self.source)[0]["id"]
        modifier_candidature(
            numero,
            chemin_db=self.source,
            portail_url="https://jobs.agentik.co",
            portail_identifiant="thomas",
            portail_mdp="super-secret",
        )
        chemin = self._exporter()
        wb = openpyxl.load_workbook(chemin)
        for feuille in wb.worksheets:
            for ligne in feuille.iter_rows(values_only=True):
                for cellule in ligne:
                    if isinstance(cellule, str):
                        self.assertNotIn("super-secret", cellule)
                        self.assertNotIn("Portail", cellule)


if __name__ == "__main__":
    unittest.main()
