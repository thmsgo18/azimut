"""Tests de l'export Excel : structure des 4 onglets, listes déroulantes,
mise en forme conditionnelle, formules du tableau de bord et liens entre onglets."""

import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl

import db
from candidatures import ajouter_candidature
from contacts import ajouter_contact
from entreprises import ajouter_ou_recuperer_entreprise
from export_excel import exporter_excel


class TestExportExcel(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.dossier = tempfile.TemporaryDirectory()
        cls.chemin_db = str(Path(cls.dossier.name) / "test.db")
        db.initialiser_base(cls.chemin_db)
        ajouter_ou_recuperer_entreprise(
            "AgentikCo",
            site_web="https://agentik.co",
            contexte_actus="Startup agents IA, série A en 2026.",
            chemin_db=cls.chemin_db,
        )
        ajouter_candidature(
            "AgentikCo",
            "Stage agents IA",
            date_envoi="2026-08-20",
            statut="Envoyée",
            priorite="Haute",
            sous_domaine="Orchestration multi-agents",
            lien_offre="https://agentik.co/jobs/stage",
            texte_offre="Concevoir des agents multi-étapes.",
            gratification=1400,
            ville="Paris",
            mode_travail="Hybride",
            source="LinkedIn",
            chemin_db=cls.chemin_db,
        )
        ajouter_candidature(
            "Mistral AI", "Stage RAG", statut="À préparer", chemin_db=cls.chemin_db
        )
        ajouter_contact(
            "AgentikCo",
            "Marie Petit",
            poste="Lead AI",
            email="marie@agentik.co",
            chemin_db=cls.chemin_db,
        )
        cls.chemin_xlsx = str(Path(cls.dossier.name) / "export.xlsx")
        exporter_excel(cls.chemin_xlsx, chemin_db=cls.chemin_db)
        cls.wb = openpyxl.load_workbook(cls.chemin_xlsx)

    @classmethod
    def tearDownClass(cls):
        cls.dossier.cleanup()

    def test_quatre_onglets(self):
        self.assertEqual(
            self.wb.sheetnames,
            ["Suivi candidatures", "Entreprises", "Contacts", "Tableau de bord"],
        )

    def test_onglet_suivi_contenu(self):
        ws = self.wb["Suivi candidatures"]
        self.assertEqual(ws["A1"].value, "Entreprise")
        self.assertEqual(ws["I1"].value, "Statut")
        # Ligne 2 = exemple jaune italique, données à partir de la ligne 3.
        self.assertTrue(ws["A2"].font.italic)
        self.assertEqual(ws["A2"].fill.start_color.rgb, "FFFFF2CC")
        self.assertEqual(ws.freeze_panes, "A3")
        # Tri par date d'envoi décroissante : AgentikCo (datée) avant Mistral (sans date)...
        # les candidatures sans date passent en dernier.
        self.assertEqual(ws["A3"].value, "AgentikCo")
        self.assertEqual(ws["B3"].value, "20/08/2026")
        self.assertEqual(ws["I3"].value, "Envoyée")
        self.assertEqual(ws["P3"].value, 1400)
        self.assertEqual(ws["A4"].value, "Mistral AI")

    def test_liens_hyperlink_match(self):
        ws = self.wb["Suivi candidatures"]
        self.assertIn("HYPERLINK", ws["W3"].value)
        self.assertIn("MATCH($A3,Entreprises!$A:$A,0)", ws["W3"].value)
        self.assertIn("MATCH($A3,Contacts!$A:$A,0)", ws["X3"].value)
        ws_ent = self.wb["Entreprises"]
        self.assertIn("MATCH($A2,'Suivi candidatures'!$A:$A,0)", ws_ent["F2"].value)
        ws_contacts = self.wb["Contacts"]
        self.assertIn("MATCH($A2,'Suivi candidatures'!$A:$A,0)", ws_contacts["L2"].value)

    def test_listes_deroulantes(self):
        ws = self.wb["Suivi candidatures"]
        validations = {str(dv.sqref): dv.formula1 for dv in ws.data_validations.dataValidation}
        self.assertEqual(len(validations), 7)
        formules = "\n".join(validations.values())
        self.assertIn("À préparer,Envoyée,Relancée,Réponse reçue,Entretien,Refus,Accepté", formules)
        self.assertIn("Haute,Moyenne,Basse", formules)
        self.assertIn("Présentiel,Hybride,Full remote", formules)
        ws_contacts = self.wb["Contacts"]
        self.assertEqual(len(ws_contacts.data_validations.dataValidation), 2)

    def test_mise_en_forme_conditionnelle(self):
        ws = self.wb["Suivi candidatures"]
        regles = []
        for plage in ws.conditional_formatting:
            for regle in plage.rules:
                regles.append((str(plage.sqref), regle.formula[0]))
        plages = {p for p, _ in regles}
        self.assertTrue(any(p.startswith("I3:I") for p in plages))  # Statut
        self.assertTrue(any(p.startswith("H3:H") for p in plages))  # Priorité
        formules = {f for _, f in regles}
        self.assertIn('"Accepté"', formules)
        self.assertIn('"Haute"', formules)
        self.assertEqual(len(regles), 10)  # 7 statuts + 3 priorités

    def test_onglet_entreprises(self):
        ws = self.wb["Entreprises"]
        noms = {ws.cell(row=l, column=1).value for l in (2, 3)}
        self.assertEqual(noms, {"AgentikCo", "Mistral AI"})
        # Compteur de candidatures par formule, pas de valeur codée en dur.
        self.assertIn("COUNTIF('Suivi candidatures'!$A$3:$A$", ws["E2"].value)

    def test_onglet_contacts(self):
        ws = self.wb["Contacts"]
        self.assertEqual(ws["A2"].value, "AgentikCo")
        self.assertEqual(ws["B2"].value, "Marie Petit")
        self.assertEqual(ws["E2"].value, "marie@agentik.co")

    def test_tableau_de_bord_formules(self):
        ws = self.wb["Tableau de bord"]
        self.assertEqual(ws["A1"].value, "Tableau de bord — candidatures stage")
        # Compteurs par statut via COUNTIF (aucune valeur en dur).
        self.assertIn('=COUNTIF(', ws["B4"].value)
        self.assertIn('"À préparer"', ws["B4"].value)
        self.assertIn('"Accepté"', ws["B10"].value)
        self.assertIn("=COUNTA('Suivi candidatures'!$A$3:$A$", ws["B12"].value)
        self.assertEqual(ws["B13"].value, "=IFERROR((B7+B8+B9+B10)/B12,0)")
        self.assertEqual(ws["B13"].number_format, "0%")
        # Sections sous-domaines et contacts.
        self.assertIn('"Orchestration multi-agents"', ws["B17"].value)
        self.assertIn("Contacts!$H$2", ws["B25"].value)

    def test_export_relancable(self):
        # Relancer l'export ne perd rien : le fichier est régénéré depuis la base.
        chemin2 = str(Path(self.dossier.name) / "export2.xlsx")
        exporter_excel(chemin2, chemin_db=self.chemin_db)
        wb2 = openpyxl.load_workbook(chemin2)
        self.assertEqual(wb2["Suivi candidatures"]["A3"].value, "AgentikCo")


if __name__ == "__main__":
    unittest.main()
